"""Create a reviewable snapshot of the JSA Occupation Shortage List (OSCA 2024).

The JSA categories are location classifications, not a 1–5 severity scale:
S = Shortage, M = Metropolitan shortage, R = Regional shortage, NS = No shortage.

Usage:
  python3 scripts/import-au-jsa-osl.py --input "/path/to/2025 Occupation Shortage List - 6 digit ANZSCO and OSCA.xlsx"
"""
from __future__ import annotations

import argparse
import hashlib
import json
import ssl
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from urllib.request import urlopen

import certifi
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "src/data/au-jsa-osl-2025.json"
SOURCE_URL = "https://www.jobsandskills.gov.au/sites/default/files/2025-10/2025%20Occupation%20Shortage%20List%20-%206%20digit%20ANZSCO%20and%20OSCA.xlsx"
STATES = ("NSW", "VIC", "QLD", "SA", "WA", "TAS", "NT", "ACT")
VALID_RATINGS = {"NS", "S", "R", "M"}


def download() -> bytes:
    context = ssl.create_default_context(cafile=certifi.where())
    with urlopen(SOURCE_URL, context=context) as response:  # nosec B310 - fixed official government URL
        return response.read()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, help="A previously downloaded official JSA workbook")
    args = parser.parse_args()

    content = args.input.read_bytes() if args.input else download()
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook["2025 OSL (OSCA 2024)"]
    if sheet["A8"].value != "Occupation code (OSCA 2024)":
        raise RuntimeError("Unexpected JSA OSL OSCA worksheet header")

    occupations = []
    for row in sheet.iter_rows(min_row=9, values_only=True):
        if not isinstance(row[0], int):
            continue
        national_rating = row[2]
        state_ratings = dict(zip(STATES, row[3:11]))
        if national_rating not in VALID_RATINGS or any(rating not in VALID_RATINGS for rating in state_ratings.values()):
            raise RuntimeError(f"Unexpected JSA rating for OSCA {row[0]}")
        occupations.append({
            "oscaCode": str(row[0]).zfill(6),
            "title": row[1],
            "nationalRating": national_rating,
            "stateRatings": state_ratings,
        })

    payload = {
        "source": {
            "name": "Jobs and Skills Australia 2025 Occupation Shortage List",
            "url": SOURCE_URL,
            "year": 2025,
            "retrievedAt": datetime.now(timezone.utc).isoformat(),
            "contentHash": hashlib.sha256(content).hexdigest(),
        },
        "ratingDefinitions": {
            "S": "Shortage",
            "M": "Metropolitan shortage",
            "R": "Regional shortage",
            "NS": "No shortage",
        },
        "occupations": occupations,
    }
    OUTPUT.write_text(json.dumps(payload, indent=2) + "\n")
    print(f"[au-jsa-osl] wrote {len(occupations)} OSCA occupations to {OUTPUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
