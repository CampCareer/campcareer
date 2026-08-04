"""AU parser: ASGS 2021 Significant Urban Areas + States/Territories, ABS.

SUA populations come from ABS 3218.0 Table 2 (ERP at 30 June 2025); state and
territory populations from ABS 3101.0 Table 8 (persons, all ages, 30 June
2025). Boundaries are the ASGS 2021 SUA/STE MapServer exports. SUAs are joined
to their parent state/territory by point-in-polygon of the SUA centroid
(cross-border SUAs fall to the state containing their centroid).
"""

import json

from shapely.geometry import shape as shapely_shape
from shapely.ops import unary_union

from ._common import entity, join_metric, stage_path, to_int

from geo import centroid_of, nearest_polygon, point_in_polygon

POP_REF = "2025-06-30"

STE_ORDER = ["1", "2", "3", "4", "5", "6", "7", "8", "9"]


def _read_geo(raw_dir, filename, code_field):
    data = json.load(open(stage_path(raw_dir, filename)))
    out = {}
    for f in data["features"]:
        out[f["properties"][code_field]] = f
    return out


def parse(raw_dir):
    import openpyxl
    wb = openpyxl.load_workbook(
        stage_path(raw_dir, "au_32180DS0004_2001-25.xlsx"), read_only=True,
        data_only=True)
    ws = wb["Table 2"]
    rows = list(ws.iter_rows(values_only=True))
    year_idx = None
    for i, v in enumerate(rows[4]):
        if v is not None and str(v).strip() == "2025":
            year_idx = i
    sua_pop = {}
    for r in rows[6:]:
        if r[0] is None or r[1] is None:
            continue
        try:
            code = str(int(r[0]))
        except (ValueError, TypeError):
            continue
        if r[year_idx] is None:
            continue
        sua_pop[code] = (str(r[1]).strip(), to_int(r[year_idx]))
    wb.close()

    wb = openpyxl.load_workbook(
        stage_path(raw_dir, "au_31010do002_202512.xlsx"), read_only=True,
        data_only=True)
    ws = wb["Table_8"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(v).strip() if v else "" for v in rows[4]]
    ste_col = {c: i for i, c in enumerate(hdr) if c and not c.isdigit()
               and c != "Age (years)"}
    state_cols = {}
    all_ages = None
    in_persons = False
    for r in rows:
        a0 = str(r[0]).strip() if r[0] is not None else ""
        if a0 == "PERSONS":
            in_persons = True
        if a0 == "All ages" and in_persons:
            all_ages = r
            break
    state_pop = {}
    for name, i in ste_col.items():
        if name != "Australia" and all_ages[i] is not None:
            state_pop[name] = to_int(all_ages[i])
    australia_pop = to_int(all_ages[ste_col["Australia"]])
    wb.close()

    sua_geo = _read_geo(raw_dir, "au_sua_asgs2021.geojson", "sua_code_2021")
    ste_geo = _read_geo(raw_dir, "au_ste_asgs2021.geojson", "state_code_2021")

    ste_polys = {}
    for code in STE_ORDER:
        f = ste_geo.get(code)
        if f is not None:
            ste_polys[code] = shapely_shape(f["geometry"])

    # --- regions (states and territories) ---
    regions = []
    region_by_code = {}
    for code in STE_ORDER:
        f = ste_geo.get(code)
        if f is None:
            continue
        p = f["properties"]
        name = p["state_name_2021"]
        geom = shapely_shape(f["geometry"])
        c = centroid_of(geom)
        pop = state_pop.get(name)
        if code == "9" and pop is None:
            pop = australia_pop - sum(state_pop.values())
        regions.append(entity(
            "region", "au-abs-31010do002-v1", "AU:ste:%s" % code,
            name, name, pop, POP_REF, code,
            "state/territory (STE)",
            coordinates=({"latitude": c[1], "longitude": c[0]}
                         if c is not None else None),
            coordinate_derivation="polygon_centroid",
            area_km2=p.get("area_albers_sqkm"),
            parent_region_code="AU", parent_region_name="Australia",
            parent_join_method="direct"))
        region_by_code[code] = regions[-1]

    # --- cities (significant urban areas) ---
    ste_list = [ste_polys[c] for c in STE_ORDER if c in ste_polys]
    ste_ids = [c for c in STE_ORDER if c in ste_polys]
    cities = []
    pip = 0
    nearest = 0
    pop_joined = 0
    for code, f in sua_geo.items():
        name = f["properties"]["sua_name_2021"]
        if name.startswith("Not in any") or name.startswith("Rest of") or \
                name.startswith("Outside"):
            continue
        pop = sua_pop.get(code)
        if pop is None:
            continue
        pop_joined += 1
        geom = shapely_shape(f["geometry"])
        c = centroid_of(geom)
        ste_code = None
        method = None
        if c is not None:
            for i, g in enumerate(ste_list):
                if point_in_polygon(c[0], c[1], g):
                    ste_code = ste_ids[i]
                    method = "point_in_polygon"
                    pip += 1
                    break
        if ste_code is None and c is not None:
            idx = nearest_polygon(c[0], c[1], ste_list)
            if idx is not None:
                ste_code = ste_ids[idx]
                method = "nearest_ste"
                nearest += 1
        reg = region_by_code.get(ste_code) if ste_code else None
        cities.append(entity(
            "city", "au-abs-32180ds0004-v1", "AU:sua:%s" % code,
            name, name, pop[1], POP_REF, code,
            "significant urban area (SUA)",
            coordinates=({"latitude": c[1], "longitude": c[0]}
                         if c is not None else None),
            coordinate_derivation="polygon_centroid",
            area_km2=f["properties"].get("area_albers_sqkm"),
            parent_region_code=ste_code,
            parent_region_name=reg["short_name"] if reg else None,
            parent_join_method=method))

    # --- country ---
    aus_union = unary_union([ste_polys[c] for c in ste_polys])
    c = centroid_of(aus_union)
    country = entity(
        "country", "au-abs-31010do002-v1", "AU:country",
        "Australia", "Australia", australia_pop, POP_REF, None, "country",
        coordinates={"latitude": c[1], "longitude": c[0]},
        coordinate_derivation="polygon_centroid",
        area_km2=sum(f["properties"].get("area_albers_sqkm")
                     for f in ste_geo.values()
                     if f["properties"]["state_code_2021"] in STE_ORDER),
        parent_region_code=None, parent_region_name=None,
        parent_join_method=None)

    return {
        "country": country,
        "regions": regions,
        "cities": cities,
        "join_metrics": [
            join_metric("sua_population_join",
                        len([c for c in sua_geo
                             if not sua_geo[c]["properties"]["sua_name_2021"]
                             .startswith(("Not in any", "Rest of", "Outside"))]),
                        pop_joined, "code_join"),
            join_metric("sua_to_ste_point_in_polygon", len(cities), pip,
                        "point_in_polygon"),
            join_metric("sua_to_ste_nearest", len(cities) - pip, nearest,
                        "nearest_ste"),
        ],
        "notes": [
            "City universe is the 102 ASGS 2021 Significant Urban Areas with "
            "ERP at 30 June 2025 (cross-border SUAs appear once under their "
            "centroid state).",
            "State/territory population is ABS 3101.0 persons all ages at "
            "30 June 2025; Other Territories = Australia minus the 8 states "
            "and territories.",
            "SUA areas are the ASGS Albers equal-area square-kilometres.",
        ],
    }
