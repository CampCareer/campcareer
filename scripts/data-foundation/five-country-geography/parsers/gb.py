"""GB parser: England & Wales towns (2019) + English regions, ONS.

City universe is the ONS "Understanding Towns and Cities in England and Wales"
dataset (1,186 towns + 53 cities), which uses a mix of BUA and BUASD 2011
codes; each code resolves to a polygon in the BUA or BUASD boundary file.
Region populations come from the MYE2 mid-2022 estimates; the country total
(United Kingdom) is 67,596,281.
"""

import json

from shapely.geometry import shape as shapely_shape
from shapely.ops import unary_union

from ._common import entity, join_metric, stage_path, to_int

from geo import bng_to_wgs84, centroid_of

POP_REF = "2019-06-30"
REGION_POP_REF = "2022-06-30"

REGION_BY_LABEL = {
    "North East": "E12000001",
    "North West": "E12000002",
    "Yorkshire and The Humber": "E12000003",
    "East Midlands": "E12000004",
    "West Midlands": "E12000005",
    "East of England": "E12000006",
    "London": "E12000007",
    "South East": "E12000008",
    "South West": "E12000009",
    "Wales": "W92000004",
}

REGION_NAMES = {
    "E12000001": "North East",
    "E12000002": "North West",
    "E12000003": "Yorkshire and The Humber",
    "E12000004": "East Midlands",
    "E12000005": "West Midlands",
    "E12000006": "East of England",
    "E12000007": "London",
    "E12000008": "South East",
    "E12000009": "South West",
}


def _strip_suffix(name):
    name = name.strip()
    for suffix in (" BUASD", " BUA"):
        if name.endswith(suffix):
            return name[: -len(suffix)], suffix[1:]
    return name, None


def _load_towns(raw_dir):
    import openpyxl
    wb = openpyxl.load_workbook(
        stage_path(raw_dir, "gb_towns_datadownloadv2.xlsx"), read_only=True)
    towns = {}
    for sheet in ["Towns (5,000 to 225,000)", "Cities (>225,000) and London"]:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(v).strip() if v else "" for v in rows[0]]
        ci, ni, ri, yi = (hdr.index("TOWN_2011CODE"),
                          hdr.index("TOWN_2011NAME"),
                          hdr.index("REGION/COUNTRY"), hdr.index("2019"))
        for r in rows[1:]:
            if r[ci] is None:
                continue
            age = str(r[7] if sheet.startswith("Towns") else r[4]).strip()
            if age != "TOTAL":
                continue
            code = str(r[ci]).strip()
            towns[code] = {
                "name": str(r[ni]).strip(),
                "region": str(r[ri]).strip(),
                "pop": int(r[yi]),
            }
    wb.close()
    return towns


def _load_geometries(raw_dir, filename, code_field):
    data = json.load(open(stage_path(raw_dir, filename)))
    out = {}
    for f in data["features"]:
        code = f["properties"][code_field]
        out[code] = shapely_shape(f["geometry"])
    return out


def parse(raw_dir):
    towns = _load_towns(raw_dir)
    bua = _load_geometries(raw_dir, "gb_bua11.geojson", "BUA11CD")
    buasd = _load_geometries(raw_dir, "gb_buasd11.geojson", "BUASD11CD")
    countries_geo = _load_geometries(
        raw_dir, "gb_countries_2022_geometry.geojson", "CTRY22CD")

    import openpyxl
    wb = openpyxl.load_workbook(
        stage_path(raw_dir, "gb_mye22final.xlsx"), read_only=True)
    ws = wb["MYE2 - Persons"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(v).strip() if v else "" for v in rows[7]]
    ci, gi, ai = hdr.index("Code"), hdr.index("Geography"), hdr.index("All ages")
    region_pop = {}
    country_pop = None
    for r in rows[8:]:
        code = str(r[ci]).strip()
        if code == "K02000001":
            country_pop = to_int(r[ai])
        if code.startswith("E1200") and str(r[gi]).strip() == "Region":
            region_pop[code] = to_int(r[ai])
    wb.close()

    regions = []
    region_by_code = {}
    regions_props = {}
    for f in json.load(open(stage_path(raw_dir, "gb_regions_2022_boundaries.geojson")))["features"]:
        regions_props[f["properties"]["RGN22CD"]] = f["properties"]
    for code, name in REGION_NAMES.items():
        props = regions_props.get(code, {})
        lon, lat = props.get("LONG"), props.get("LAT")
        coords = ({"latitude": lat, "longitude": lon}
                  if lon is not None and lat is not None else None)
        regions.append(entity(
            "region", "gb-ons-mye22-v1", "GB:region:%s" % code,
            name, name, region_pop.get(code), REGION_POP_REF, code,
            "region (E1200)",
            coordinates=coords,
            coordinate_derivation="official_centroid_property",
            area_km2=None,
            parent_region_code="E92000001", parent_region_name="England",
            parent_join_method="direct"))
        region_by_code[code] = regions[-1]

    cities = []
    geometry_joined = 0
    for code, t in towns.items():
        geom = bua.get(code) or buasd.get(code)
        name, suffix = _strip_suffix(t["name"])
        c = centroid_of(geom)
        coords = None
        if c is not None:
            lon, lat = bng_to_wgs84(c[0], c[1])
            coords = {"latitude": lat, "longitude": lon}
            geometry_joined += 1
        level = ("built-up area sub-division (BUASD)" if suffix == "BUASD"
                 else "built-up area (BUA)")
        rgn_code = REGION_BY_LABEL.get(t["region"])
        reg = region_by_code.get(rgn_code)
        cities.append(entity(
            "city", "gb-ons-understanding-towns-v1", "GB:town:%s" % code,
            name, name, t["pop"], POP_REF, code, level,
            coordinates=coords,
            coordinate_derivation="polygon_centroid_transformed",
            area_km2=None,
            parent_region_code=rgn_code,
            parent_region_name=reg["short_name"] if reg else t["region"],
            parent_join_method="direct"))

    gb_union = unary_union(list(countries_geo.values()))
    cc = centroid_of(gb_union)
    lon, lat = cc[0], cc[1]
    country = entity(
        "country", "gb-ons-mye22-v1", "GB:country",
        "United Kingdom", "United Kingdom", country_pop, REGION_POP_REF,
        None, "country",
        coordinates={"latitude": lat, "longitude": lon},
        coordinate_derivation="polygon_centroid",
        area_km2=None, parent_region_code=None, parent_region_name=None,
        parent_join_method=None)

    return {
        "country": country,
        "regions": regions,
        "cities": cities,
        "join_metrics": [
            join_metric("town_geometry_join", len(towns), geometry_joined,
                        "code_join_bua_buasd"),
            join_metric("region_population_join", len(REGION_NAMES),
                        sum(1 for c in REGION_NAMES if c in region_pop),
                        "code_join"),
            join_metric("town_region_join", len(towns),
                        sum(1 for t in towns.values()
                            if t["region"] in REGION_BY_LABEL), "label_join"),
        ],
        "notes": [
            "Towns dataset uses a mix of BUA (E3400/W3800) and BUASD "
            "(E3500/W3700/K0600) 2011 codes; all 1,239 codes resolve to a "
            "boundary polygon.",
            "City population is the 2019 TOTAL row; region/country population "
            "is mid-2022 from the MYE2 estimates.",
            "The dataset covers England and Wales; English towns are assigned "
            "to the 9 English regions, Welsh towns to Wales (W92000004).",
        ],
    }
