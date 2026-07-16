"""Build the OSCA occupations that map to the Home Affairs Core Skills Occupation List.

The output is a reviewable source snapshot. It does not write to Supabase.
Run `npx tsx scripts/seed-au-csol-priority-occupations.ts` after reviewing it.
"""
from __future__ import annotations

import hashlib
import json
import ssl
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from urllib.request import urlopen

from openpyxl import load_workbook
from pypdf import PdfReader
import certifi

ROOT = Path(__file__).resolve().parents[1]
OSCA_PROFILES = ROOT / "src/data/au-osca-occupation-profiles.json"
OUTPUT = ROOT / "src/data/au-csol-priority-occupations.json"
CSOL_URL = "https://immi.homeaffairs.gov.au/Documents/core-sol.pdf"
CORRESPONDENCE_URL = "https://www.abs.gov.au/statistics/classifications/osca-occupation-standard-classification-australia/2024-version-1-0/data-downloads/OSCA%20correspondence%20tables%20v2.xlsx"


def download(url: str) -> bytes:
    context = ssl.create_default_context(cafile=certifi.where())
    with urlopen(url, context=context) as response:  # nosec B310 - fixed official government URLs
        return response.read()


def six_digit_codes_from_pdf(content: bytes) -> set[str]:
    text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(content)).pages)
    import re
    return set(re.findall(r"\b\d{6}\b", text))


def osca_to_anzsco(content: bytes) -> dict[str, set[str]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook["Table 2"]
    if sheet["A5"].value != "OSCA 2024 v1.0" or sheet["C5"].value != "ANZSCO v1.3":
        raise RuntimeError("Unexpected ABS correspondence Table 2 header")

    result: dict[str, set[str]] = {}
    for row in sheet.iter_rows(min_row=6, values_only=True):
        osca = str(row[0] or "").zfill(6)
        anzsco = str(row[2] or "").zfill(6)
        if osca.isdigit() and len(osca) == 6 and anzsco.isdigit() and len(anzsco) == 6:
            result.setdefault(osca, set()).add(anzsco)
    return result


def main() -> None:
    csol_bytes = download(CSOL_URL)
    correspondence_bytes = download(CORRESPONDENCE_URL)
    csol_codes = six_digit_codes_from_pdf(csol_bytes)
    mapping = osca_to_anzsco(correspondence_bytes)
    profiles = json.loads(OSCA_PROFILES.read_text())["occupations"]
    selected = [
        {
            "oscaCode": profile["code"],
            "anzscoV13Codes": sorted(mapping.get(profile["code"], [])),
            "title": profile["title"],
            "skillLevel": profile["skillLevel"],
            "registrationOrLicensing": profile["registrationOrLicensing"],
        }
        for profile in profiles
        if set(mapping.get(profile["code"], [])) & csol_codes
    ]
    selected.sort(key=lambda item: item["oscaCode"])
    payload = {
        "source": {
            "name": "Department of Home Affairs Core Skills Occupation List matched via ABS OSCA 2024 v1.0 correspondence Table 2",
            "csolUrl": CSOL_URL,
            "correspondenceUrl": CORRESPONDENCE_URL,
            "retrievedAt": datetime.now(timezone.utc).isoformat(),
            "csolContentHash": hashlib.sha256(csol_bytes).hexdigest(),
            "correspondenceContentHash": hashlib.sha256(correspondence_bytes).hexdigest(),
        },
        "csolAnzscoV13Codes": sorted(csol_codes),
        "occupations": selected,
    }
    OUTPUT.write_text(json.dumps(payload, indent=2) + "\n")
    print(f"[au-csol] wrote {len(selected)} OSCA occupations; CSOL contains {len(csol_codes)} ANZSCO codes.")


if __name__ == "__main__":
    main()
